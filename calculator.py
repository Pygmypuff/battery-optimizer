import copy
from dataclasses import dataclass
from enum import Enum
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.colors import ColorChoice, SchemeColor
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import (
    RichTextProperties, ListStyle, Paragraph,
    ParagraphProperties, CharacterProperties
)
import pandas as pd
from pathlib import Path
from datetime import datetime, time, date
import math
from nordpool import elspot
import pytz
import bisect
import shutil
import os
import re
import zipfile


from app_paths import bundle_dir, output_dir as default_output_dir
from battery_optimizer import (
    BatteryAction,
    StationConfig,
    StationState,
    compute_charge_discharge_ratio,
    optimise_battery_schedule,
    print_schedule,
    rerun_for_remaining_day,
)
from station_config import calculate_usable_battery_capacity, load_config

# Config values (StationConfig fields, total_battery_capacity,
# bottom_unusable_pct, red_line_threshold) now live in station_config.py /
# config.json, editable from the GUI's config window. This module-level
# snapshot is only a fallback default for callers that don't go through
# run_nordpool() (e.g. direct StationConfig-needing calls); run_nordpool()
# itself always reloads fresh so GUI-saved changes apply on the next run
# without restarting the app.
_DEFAULT_APP_CFG = load_config()
BASE_CFG = _DEFAULT_APP_CFG.station_config()

# OTHER CONFIG VALUES
total_battery_capacity = _DEFAULT_APP_CFG.total_battery_capacity
bottom_unusable_pct = _DEFAULT_APP_CFG.bottom_unusable_pct
top_unusable_pct = 5 # 5% of battery capacity is unusable (top) — not currently used below
red_line_threshold = _DEFAULT_APP_CFG.red_line_threshold

# Where excel_template.xlsx (a bundled, read-only resource) lives. Uses
# bundle_dir() rather than raw __file__ so this still resolves correctly
# once frozen by PyInstaller — see app_paths.py.
SCRIPT_DIR = str(bundle_dir())

# --- Domain types ---

class BatteryAction(str, Enum):
    CHARGE    = "CHARGE"
    DISCHARGE = "DISCHARGE"
    HOLD      = "HOLD"

@dataclass
class SlotResult:
    """Outcome for a single 15-minute slot."""
    slot_index:        int
    action:            BatteryAction
    energy_charged:    float   # MWh added to battery         (0 unless CHARGE)
    energy_discharged: float   # MWh drawn from battery       (0 unless DISCHARGE)
    power_sold:        float   # MW sold to grid this slot
    revenue:           float   # EUR earned this slot
    battery_level_end: float   # MWh in battery at end of slot

COLOR_MAP = {
    BatteryAction.CHARGE:    "4169E1",  # blue
    BatteryAction.DISCHARGE: "ED7D31",  # orange
    BatteryAction.HOLD:      "828481",  # grey
}

def datetime_to_slot_index(dt: datetime, slots: list[datetime]) -> int:
    """
    Return the index in *slots* of the first datetime >= *dt*.

    - If *dt* matches a datetime in *slots* exactly, return that index.
    - If *dt* falls between two entries, return the index of the next one.
    - Raises ValueError if *dt* is after the last entry in *slots*.
    """
    index = bisect.bisect_left(slots, dt)

    if index >= len(slots):
        raise ValueError(
            f"{dt} is after the last slot ({slots[-1]}); no valid slot exists."
        )

    return index


def battery_optimizer_run(prices: list[float], state: StationState, slots_elapsed: int, cfg: StationConfig):
    """
    Run the battery optimizer and return the result, starting from the given state and slot index.
    """
    print("=" * 78)
    print(f"  P = {state.station_power:.4f} MW  |  "
          f"Charge:discharge ratio X (informational only, not a solver "
          f"constraint) = {compute_charge_discharge_ratio(state.station_power, cfg):.4f}  |  "
          f"Battery = {state.battery_level:.4f} MWh")
    print("=" * 78)

    result = rerun_for_remaining_day(
            remaining_prices = prices[slots_elapsed:],
            cfg              = cfg,
            updated_state    = state,
            slots_elapsed    = slots_elapsed,
        )
    return result


def fetch_prices():
    """
    Fetch electricity prices from nordpool for today and tomorrow,
    Converts prices to Riga timezone and filters out prices before 14:00 today.
    Returns: dictionary with "datetime" and "price" lists.
    throws: ValueError if tomorrow's prices cannot be fetched (e.g. not available yet).
    """
    riga_tz = pytz.timezone('Europe/Riga')
    now = datetime.now(riga_tz)
    target = now.replace(hour=14, minute=0, second=0, microsecond=0) # 14:00 today

    prices_spot = elspot.Prices()
    today_prices = prices_spot.fetch( # get todays prices
        end_date=date.today(),
        areas=["LV"],
        resolution=15,
    )
    tomorrow_prices = prices_spot.fetch( # get tomorrows prices
        areas=["LV"],
        resolution=15,
    )

    # Check if prices were successfully fetched
    if not tomorrow_prices:
        raise ValueError("Failed to fetch tomorrow's prices. Prices might not be available yet.")
    
    prices_list = {
    "datetime": [],
    "price": [],
    }

    # add prices to prices_list, converting timestamps to Riga timezone
    for entry in today_prices["areas"]["LV"]["values"]:
        prices_list["datetime"].append(entry["start"].astimezone(riga_tz))
        prices_list["price"].append(entry["value"])

    for entry in tomorrow_prices["areas"]["LV"]["values"]:
        prices_list["datetime"].append(entry["start"].astimezone(riga_tz))
        prices_list["price"].append(entry["value"])

    # Filter out prices that are before the target time
    prices_list["datetime"], prices_list["price"] = zip(*[
        (dt, price) for dt, price in zip(prices_list["datetime"], prices_list["price"])
        if dt >= target
    ])

    return prices_list


def duplicate_excel(src_path: str, dst_path: str) -> None:
    #duplicate an excel file, preserving all formatting and charts
    shutil.copy2(src_path, dst_path)

def replace_column_c(filepath: str, values: list[float]) -> None:
    """
    Replace values in column C (starting from C2) with the provided list of values.
    """
    if len(values) != 140:
        raise ValueError("Expected 140 values, got {len(values)}")

    wb = load_workbook(filepath)
    ws = wb["tabula"]

    for i, value in enumerate(values):
        ws.cell(row=i + 2, column=3, value=value)

    wb.save(filepath)

def color_chart_bars(filepath: str, slots: list[SlotResult], start_index: int) -> None:
    """
    Colors the bars in the excel chart to match the slot actions.
    """
    if len(slots) + start_index != 140:
        raise ValueError(f"Expected {140 - start_index} SlotResults, got {len(slots)}")

    wb = load_workbook(filepath)
    ws = wb["tabula"]
    chart = ws._charts[0]
    ser = chart.series[0]

    dpt_by_idx = {pt.idx: pt for pt in ser.dPt}

    for i, slot in enumerate(slots):
        bar_index = start_index + i
        pt = dpt_by_idx[bar_index]
        pt.spPr.solidFill.schemeClr = None
        pt.spPr.solidFill.srgbClr = COLOR_MAP[slot.action]

    wb.save(filepath)

"""
Helper functions to create a RichText object with the specified hex color for the text.
"""
def _make_txPr(hex_color: str) -> RichText:
    rPr = CharacterProperties(sz=1000.0, b=True)
    rPr.solidFill = ColorChoice()
    rPr.solidFill.srgbClr = hex_color
    pPr = ParagraphProperties(defRPr=rPr)
    para = Paragraph(pPr=pPr, endParaRPr=CharacterProperties())
    bodyPr = RichTextProperties(rot=-5400000, anchor="ctr", anchorCtr=True)
    return RichText(bodyPr=bodyPr, lstStyle=ListStyle(), p=[para])

def _make_txPr_default() -> RichText:
    solidFill = ColorChoice()
    solidFill.schemeClr = SchemeColor(val="tx1")
    rPr = CharacterProperties(sz=1000.0, b=True)
    rPr.solidFill = solidFill
    pPr = ParagraphProperties(defRPr=rPr)
    para = Paragraph(pPr=pPr, endParaRPr=CharacterProperties())
    bodyPr = RichTextProperties(rot=-5400000, anchor="ctr", anchorCtr=True)
    return RichText(bodyPr=bodyPr, lstStyle=ListStyle(), p=[para])


def color_label_text_below_threshold(
    filepath: str,
    values: list[float],
    threshold: float,
    sheet_name: str = "tabula",
    chart_index: int = 0,
    series_index: int = 0,
) -> None:
    """
    Colors the data labels in the chart red if they are below red-line-threshold.
    """
    if len(values) != 140:
        raise ValueError(f"Expected 140 values, got {len(values)}")

    wb = load_workbook(filepath)
    ws = wb[sheet_name]
    chart = ws._charts[chart_index]
    ser = chart.series[series_index]

    for lbl in ser.dLbls.dLbl:
        lbl.txPr = _make_txPr("FF0000") if values[lbl.idx] < threshold else _make_txPr_default()

    wb.save(filepath)

def apply_chart_background(filepath: str) -> None:
    """
    Injects a gradient background into the chart XML to visually distinguish between 24h segments.
 
    IMPORTANT: Call this AFTER all openpyxl wb.save() calls
    """
    # Gradient stop positions (units: 1/1000 of a percent, 0–100 000)
    s1_end = round(40 / 140 * 100_000)   # 28571 — end of first white band
    s2_end = round(136 / 140 * 100_000)  # 97143 — end of grey band
 
    # Read all files from the xlsx zip
    files: dict[str, bytes] = {}
    with zipfile.ZipFile(filepath, "r") as z:
        for name in z.namelist():
            files[name] = z.read(name)
 
    chart_xml = files["xl/charts/chart1.xml"].decode("utf-8")
 
    # openpyxl strips namespace prefixes on save (e.g. <c:plotArea> → <plotArea>)
    # so we detect which form is present and adapt accordingly.
    if "</c:plotArea>" in chart_xml:
        close_tag = "</c:plotArea>"
        open_spPr = (
            '<c:spPr'
            ' xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart"'
            ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        )
        close_spPr = "</c:spPr>"
    else:
        close_tag = "</plotArea>"
        open_spPr = '<spPr xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        close_spPr = "</spPr>"
 
    gradient_spPr = (
        open_spPr
        + "<a:gradFill><a:gsLst>"
        + '<a:gs pos="0"><a:srgbClr val="FFFFFF"/></a:gs>'
        + f'<a:gs pos="{s1_end}"><a:srgbClr val="FFFFFF"/></a:gs>'
        + f'<a:gs pos="{s1_end + 1}"><a:srgbClr val="E0E0E0"/></a:gs>'
        + f'<a:gs pos="{s2_end}"><a:srgbClr val="E0E0E0"/></a:gs>'
        + f'<a:gs pos="{s2_end + 1}"><a:srgbClr val="FFFFFF"/></a:gs>'
        + '<a:gs pos="100000"><a:srgbClr val="FFFFFF"/></a:gs>'
        + "</a:gsLst>"
        + '<a:lin ang="0" scaled="0"/>'
        + "</a:gradFill>"
        + "<a:ln><a:noFill/></a:ln>"
        + close_spPr
    )
 
    # Remove any previously injected plotArea spPr (makes this call idempotent)
    for open_pat, close_pat in [
        (r"<c:spPr[^>]*>", r"</c:spPr>\s*</c:plotArea>"),
        (r"<spPr[^>]*>",   r"</spPr>\s*</plotArea>"),
    ]:
        chart_xml = re.sub(
            open_pat + r".*?" + close_pat,
            close_tag,
            chart_xml,
            count=1,
            flags=re.DOTALL,
        )
 
    # Inject the gradient spPr immediately before </plotArea>
    chart_xml = chart_xml.replace(close_tag, gradient_spPr + close_tag, 1)
 
    files["xl/charts/chart1.xml"] = chart_xml.encode("utf-8")
 
    # Write back atomically
    tmp = filepath + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in files.items():
            zout.writestr(name, data)
    os.replace(tmp, filepath)








# ── Live (Nordpool) run ──────────────────────────────────────────────────────

def _run_nordpool_from(
    target: datetime,
    station_power: float,
    battery_pct: float,
    initial_charge_price: float,
    output_dir: str | Path | None,
) -> Path:
    """
    Shared implementation for run_nordpool() (from 14:00) and
    run_nordpool_from_now() (from the current moment): fetch live Nordpool
    prices, run the battery optimizer starting at `target`, and write a
    formatted output workbook. Returns the path to the generated .xlsx.

    Progress is reported via plain `print()` calls, same as the rest of
    this module — callers that want to capture it (e.g. a GUI) can redirect
    stdout around the call.
    """
    # Reload config fresh on every run so a change saved from the GUI's
    # config window takes effect on the very next run, without needing to
    # restart the app.
    app_cfg = load_config()
    cfg = app_cfg.station_config()

    # Defaults to the OS-managed per-user output folder (see app_paths.py),
    # not SCRIPT_DIR — once packaged, SCRIPT_DIR is inside the app bundle,
    # and generated workbooks would be lost every time the app updates.
    out_dir = Path(output_dir) if output_dir is not None else default_output_dir()
    out_dir.mkdir(parents=True, exist_ok=True)

    print("Fetching prices from nordpool...")
    prices_list = fetch_prices()
    print("=" * 78)
    print(f"Fetched {len(prices_list['price'])} electricity prices")

    slot_index = datetime_to_slot_index(target, prices_list["datetime"])
    print(f"Running battery optimizer from {target} (slot index {slot_index})...")

    result = battery_optimizer_run(
        prices=prices_list["price"],
        state=StationState(
            station_power=station_power,
            battery_level=calculate_usable_battery_capacity(battery_pct, app_cfg),
            initial_charge_price=initial_charge_price,
        ),
        slots_elapsed=slot_index,
        cfg=cfg,
    )
    print("Battery optimizer run complete.")
    print(f"  BATTERY SCHEDULE  —  {result.slots_optimised} slots  |  "
          f"Expected revenue: {result.total_revenue:.2f} EUR")
    print("=" * 78)

    src_filepath = os.path.join(SCRIPT_DIR, "excel_template.xlsx")
    timestamp = datetime.now(target.tzinfo).strftime("%Y%m%d_%H%M%S")
    dst_filepath = str(out_dir / f"output_{timestamp}.xlsx")

    print("Duplicating template file...")
    duplicate_excel(src_path=src_filepath, dst_path=dst_filepath)

    print("Replacing prices in output file...")
    replace_column_c(dst_filepath, prices_list["price"])

    print("Coloring chart bars in output file...")
    color_chart_bars(dst_filepath, result.schedule, start_index=slot_index)

    print("Coloring labels below threshold in output file...")
    color_label_text_below_threshold(dst_filepath, prices_list["price"], app_cfg.red_line_threshold)

    print("Applying chart background in output file...")
    apply_chart_background(dst_filepath)  # This must be called last

    print(f"Done. Output saved to {dst_filepath}")
    return Path(dst_filepath)


def run_nordpool(
    station_power: float,
    battery_pct: float = 12.0,
    initial_charge_price: float = 0.0,
    output_dir: str | Path | None = None,
) -> Path:
    """Run the optimizer for the rest of today, starting at 14:00 today."""
    riga_tz = pytz.timezone('Europe/Riga')
    now = datetime.now(riga_tz)
    target = now.replace(hour=14, minute=0, second=0, microsecond=0)  # 14:00 today
    return _run_nordpool_from(target, station_power, battery_pct, initial_charge_price, output_dir)


def run_nordpool_from_now(
    station_power: float,
    battery_pct: float = 12.0,
    initial_charge_price: float = 0.0,
    output_dir: str | Path | None = None,
) -> Path:
    """
    Run the optimizer for the rest of today, starting right now — a mid-day
    rerun, e.g. after the station's actual power output has drifted from
    what was assumed at the last 14:00 run.
    """
    riga_tz = pytz.timezone('Europe/Riga')
    now = datetime.now(riga_tz)
    return _run_nordpool_from(now, station_power, battery_pct, initial_charge_price, output_dir)


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    run_nordpool(station_power=0.310)

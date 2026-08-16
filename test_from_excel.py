"""
test_from_excel.py
===================
Test harness for the battery optimizer that sources its inputs (station
power + 15-minute price series) from a BESS forecast Excel file instead of
a live nordpool fetch. This lets you re-run the exact same optimisation
logic used in calculator.py against any historical/example day.

Input file expectations (sheet "BESS (15)" by default):
  - C3            : station power for the day. Stored in kW in this sheet
                     (confirmed against the "jaudas" sheet's 0-140 kW power
                     scale and the "Ranka(15)" sheet's C1 ~ 100-150 value),
                     so it is divided by 1000 to get MW for StationConfig/
                     StationState. Pass --power-unit mw if your sheet is
                     already in MW.
  - B4:B143       : time-of-day labels for each 15-min slot (informational
                     only, not used numerically).
  - C4:C143       : price EUR/MWh for each of the 140 slots (must be
                     exactly 140 rows, matching calculator.py's expectation).

This script intentionally does NOT import calculator.py, because
calculator.py imports the `nordpool` package at module scope purely to
fetch live prices, which isn't needed for offline testing and may not even
be installed in a test environment. Config values (StationConfig fields,
total_battery_capacity, bottom_unusable_pct, red_line_threshold) come from
station_config.py instead, which both this script and calculator.py share
— editable from the GUI's config window, with no nordpool dependency.

Optionally also writes a clean, single-sheet formatted .xlsx (schedule
table + colored bar chart matching calculator.py's chart styling: blue for
CHARGE, orange for DISCHARGE/sell, grey for HOLD, red price labels below a
threshold, and a light gradient background marking day segments) built
from scratch — it does not carry over any of the extra sheets from the
input workbook.

Usage:
    python test_from_excel.py path/to/BESS2026_08_06.xlsx
    python test_from_excel.py path/to/file.xlsx --sheet "BESS (15)" \
        --battery-pct 12 --initial-charge-price 0 \
        --out schedule.csv --xlsx-out schedule.xlsx
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import shutil
import zipfile
from dataclasses import dataclass

from openpyxl import load_workbook
from openpyxl.drawing.colors import ColorChoice, SchemeColor
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import (
    RichTextProperties, ListStyle, Paragraph,
    ParagraphProperties, CharacterProperties,
)

from battery_optimizer import (
    BatteryAction,
    StationConfig,
    StationState,
    optimise_battery_schedule,
    print_schedule,
)
from station_config import calculate_usable_battery_capacity, load_config

# Snapshot for module-level defaults (e.g. the --threshold CLI flag's
# default). process_one_file() always reloads fresh via load_config() so a
# change saved from the GUI's config window applies on the very next run.
_DEFAULT_APP_CFG = load_config()
BASE_CFG = _DEFAULT_APP_CFG.station_config()
total_battery_capacity = _DEFAULT_APP_CFG.total_battery_capacity
bottom_unusable_pct = _DEFAULT_APP_CFG.bottom_unusable_pct
red_line_threshold = _DEFAULT_APP_CFG.red_line_threshold

def _color_map(app_cfg) -> dict:
    """Built fresh from config on every run — see station_config.py's
    charge_color/discharge_color/hold_color, editable from the GUI's
    "Customize output chart" window."""
    return {
        BatteryAction.CHARGE:    app_cfg.charge_color,
        BatteryAction.DISCHARGE: app_cfg.discharge_color,
        BatteryAction.HOLD:      app_cfg.hold_color,
    }


def _make_txPr(hex_color: str) -> RichText:
    """RichText run properties used to color a chart data-label's text."""
    rPr = CharacterProperties(sz=1000.0, b=True)
    rPr.solidFill = ColorChoice()
    rPr.solidFill.srgbClr = hex_color
    pPr = ParagraphProperties(defRPr=rPr)
    para = Paragraph(pPr=pPr, endParaRPr=CharacterProperties())
    bodyPr = RichTextProperties(rot=-5400000, anchor="ctr", anchorCtr=True)
    return RichText(bodyPr=bodyPr, lstStyle=ListStyle(), p=[para])


# ---------------------------------------------------------------------------
# Excel extraction
# ---------------------------------------------------------------------------

EXPECTED_SLOTS = 140
FIRST_ROW = 4
LAST_ROW = FIRST_ROW + EXPECTED_SLOTS - 1  # 143

POWER_CELL = "C3"
TIME_COL = "B"
PRICE_COL = "C"


@dataclass
class ExcelInputs:
    station_power_mw: float
    prices: list[float]
    time_labels: list[str]


def read_inputs_from_excel(
    filepath: str,
    sheet_name: str,
    power_unit: str,
) -> ExcelInputs:
    # data_only=True to get the last-cached *values* of formula cells
    # (C4:C143 are `=SUM(NordPool!...)` formulas in the example file).
    wb = load_workbook(filepath, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}"
        )
    ws = wb[sheet_name]

    raw_power = ws[POWER_CELL].value
    if raw_power is None:
        raise ValueError(f"{POWER_CELL} (station power) is empty in sheet '{sheet_name}'.")
    station_power_mw = float(raw_power) / 1000.0 if power_unit == "kw" else float(raw_power)

    prices: list[float] = []
    time_labels: list[str] = []
    for row in range(FIRST_ROW, LAST_ROW + 1):
        price_val = ws[f"{PRICE_COL}{row}"].value
        time_val = ws[f"{TIME_COL}{row}"].value
        if price_val is None:
            raise ValueError(
                f"Missing price at {PRICE_COL}{row}. Expected {EXPECTED_SLOTS} "
                f"contiguous values in {PRICE_COL}{FIRST_ROW}:{PRICE_COL}{LAST_ROW}."
            )
        prices.append(float(price_val))
        time_labels.append(str(time_val))

    if len(prices) != EXPECTED_SLOTS:
        raise ValueError(f"Expected {EXPECTED_SLOTS} prices, got {len(prices)}.")

    return ExcelInputs(station_power_mw=station_power_mw, prices=prices, time_labels=time_labels)


# ---------------------------------------------------------------------------
# Formatted .xlsx output — this reuses the REAL excel_template.xlsx that
# calculator.py itself duplicates and fills in, so the output matches
# calculator.py's output exactly (same "tabula" sheet layout, same chart,
# same formatting). These functions are calculator.py's own
# duplicate_excel/replace_column_c/color_chart_bars/
# color_label_text_below_threshold/apply_chart_background, unchanged.
# ---------------------------------------------------------------------------

TEMPLATE_SHEET = "tabula"
TEMPLATE_FIRST_ROW = 2
TEMPLATE_ROWS = 140  # rows 2..141


def duplicate_excel(src_path: str, dst_path: str) -> None:
    shutil.copy2(src_path, dst_path)


def replace_column_c(filepath: str, values: list[float]) -> None:
    """Replace values in column C (starting from C2) with the provided list."""
    if len(values) != TEMPLATE_ROWS:
        raise ValueError(f"Expected {TEMPLATE_ROWS} values, got {len(values)}")

    wb = load_workbook(filepath)
    ws = wb[TEMPLATE_SHEET]

    for i, value in enumerate(values):
        ws.cell(row=i + TEMPLATE_FIRST_ROW, column=3, value=value)

    wb.save(filepath)


def color_chart_bars(filepath: str, slots, start_index: int, color_map: dict) -> None:
    """Colors the bars in the excel chart to match the slot actions."""
    if len(slots) + start_index != TEMPLATE_ROWS:
        raise ValueError(f"Expected {TEMPLATE_ROWS - start_index} SlotResults, got {len(slots)}")

    wb = load_workbook(filepath)
    ws = wb[TEMPLATE_SHEET]
    chart = ws._charts[0]
    ser = chart.series[0]

    dpt_by_idx = {pt.idx: pt for pt in ser.dPt}

    for i, slot in enumerate(slots):
        bar_index = start_index + i
        pt = dpt_by_idx[bar_index]
        pt.spPr.solidFill.schemeClr = None
        pt.spPr.solidFill.srgbClr = color_map[slot.action]

    wb.save(filepath)


def _make_txPr_default() -> RichText:
    solidFill = ColorChoice()
    solidFill.schemeClr = SchemeColor(val="tx1")
    rPr = CharacterProperties(sz=1000.0, b=True)
    rPr.solidFill = solidFill
    pPr = ParagraphProperties(defRPr=rPr)
    para = Paragraph(pPr=pPr, endParaRPr=CharacterProperties())
    bodyPr = RichTextProperties(rot=-5400000, anchor="ctr", anchorCtr=True)
    return RichText(bodyPr=bodyPr, lstStyle=ListStyle(), p=[para])


def color_label_text_below_threshold(
    filepath: str,
    values: list[float],
    threshold: float,
    sheet_name: str = TEMPLATE_SHEET,
    chart_index: int = 0,
    series_index: int = 0,
) -> None:
    """Colors the data labels red if they are below threshold."""
    if len(values) != TEMPLATE_ROWS:
        raise ValueError(f"Expected {TEMPLATE_ROWS} values, got {len(values)}")

    wb = load_workbook(filepath)
    ws = wb[sheet_name]
    chart = ws._charts[chart_index]
    ser = chart.series[series_index]

    for lbl in ser.dLbls.dLbl:
        lbl.txPr = _make_txPr("FF0000") if values[lbl.idx] < threshold else _make_txPr_default()

    wb.save(filepath)


def apply_chart_background(filepath: str) -> None:
    """
    Injects a gradient background into the chart XML to visually distinguish
    between 24h segments. Call this AFTER all openpyxl wb.save() calls.
    """
    s1_end = round(40 / TEMPLATE_ROWS * 100_000)
    s2_end = round(136 / TEMPLATE_ROWS * 100_000)

    files: dict[str, bytes] = {}
    with zipfile.ZipFile(filepath, "r") as z:
        for name in z.namelist():
            files[name] = z.read(name)

    chart_xml = files["xl/charts/chart1.xml"].decode("utf-8")

    if "</c:plotArea>" in chart_xml:
        close_tag = "</c:plotArea>"
        open_spPr = (
            '<c:spPr'
            ' xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart"'
            ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        )
        close_spPr = "</c:spPr>"
    else:
        close_tag = "</plotArea>"
        open_spPr = '<spPr xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        close_spPr = "</spPr>"

    gradient_spPr = (
        open_spPr
        + "<a:gradFill><a:gsLst>"
        + '<a:gs pos="0"><a:srgbClr val="FFFFFF"/></a:gs>'
        + f'<a:gs pos="{s1_end}"><a:srgbClr val="FFFFFF"/></a:gs>'
        + f'<a:gs pos="{s1_end + 1}"><a:srgbClr val="E0E0E0"/></a:gs>'
        + f'<a:gs pos="{s2_end}"><a:srgbClr val="E0E0E0"/></a:gs>'
        + f'<a:gs pos="{s2_end + 1}"><a:srgbClr val="FFFFFF"/></a:gs>'
        + '<a:gs pos="100000"><a:srgbClr val="FFFFFF"/></a:gs>'
        + "</a:gsLst>"
        + '<a:lin ang="0" scaled="0"/>'
        + "</a:gradFill>"
        + "<a:ln><a:noFill/></a:ln>"
        + close_spPr
    )

    for open_pat, close_pat in [
        (r"<c:spPr[^>]*>", r"</c:spPr>\s*</c:plotArea>"),
        (r"<spPr[^>]*>",   r"</spPr>\s*</plotArea>"),
    ]:
        chart_xml = re.sub(
            open_pat + r".*?" + close_pat,
            close_tag,
            chart_xml,
            count=1,
            flags=re.DOTALL,
        )

    chart_xml = chart_xml.replace(close_tag, gradient_spPr + close_tag, 1)

    files["xl/charts/chart1.xml"] = chart_xml.encode("utf-8")

    tmp = filepath + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in files.items():
            zout.writestr(name, data)
    os.replace(tmp, filepath)


def generate_formatted_excel(
    template_path: str,
    output_path: str,
    schedule,
    prices: list[float],
    threshold: float = red_line_threshold,
    color_map: dict | None = None,
) -> None:
    print(f"Duplicating template '{template_path}' -> '{output_path}'...")
    duplicate_excel(template_path, output_path)

    print("Writing prices into template...")
    replace_column_c(output_path, prices)

    print("Coloring chart bars by action...")
    color_chart_bars(output_path, schedule, start_index=0, color_map=color_map or _color_map(_DEFAULT_APP_CFG))

    print("Coloring price labels below threshold...")
    color_label_text_below_threshold(output_path, prices, threshold)

    print("Applying chart background bands...")
    apply_chart_background(output_path)  # must be called last


# ---------------------------------------------------------------------------
# Output helpers
# ---------------------------------------------------------------------------

def write_schedule_csv(path: str, result, prices: list[float], time_labels: list[str]) -> None:
    with open(path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(
            ["slot_index", "time_label", "price_eur_mwh", "action", "power_sold_mw",
             "energy_charged_mwh", "energy_discharged_mwh", "battery_level_end_mwh", "revenue_eur"]
        )
        for i, r in enumerate(result.schedule):
            writer.writerow([
                r.slot_index,
                time_labels[i] if i < len(time_labels) else "",
                prices[r.slot_index] if r.slot_index < len(prices) else "",
                r.action.value,
                f"{r.power_sold:.4f}",
                f"{r.energy_charged:.4f}",
                f"{r.energy_discharged:.4f}",
                f"{r.battery_level_end:.4f}",
                f"{r.revenue:.4f}",
            ])


# ---------------------------------------------------------------------------
# Per-file processing (shared by single-file and batch mode)
# ---------------------------------------------------------------------------

def process_one_file(
    excel_path: str,
    sheet: str,
    power_unit: str,
    battery_pct: float,
    initial_charge_price: float,
    threshold: float,
    template: str | None,
    csv_out: str | None,
    xlsx_out: str | None,
) -> None:
    # Reload config fresh on every run so a change saved from the GUI's
    # config window takes effect on the very next run, without needing to
    # restart the app.
    app_cfg = load_config()
    cfg = app_cfg.station_config()

    print(f"Reading inputs from '{excel_path}' (sheet '{sheet}')...")
    inputs = read_inputs_from_excel(excel_path, sheet, power_unit)
    print(f"  Station power : {inputs.station_power_mw:.4f} MW "
          f"(raw C3 value {'/1000' if power_unit == 'kw' else ''})")
    print(f"  Prices        : {len(inputs.prices)} slots, "
          f"{inputs.prices[0]:.2f}..{inputs.prices[-1]:.2f} EUR/MWh")

    battery_level = calculate_usable_battery_capacity(battery_pct, app_cfg)
    state = StationState(
        station_power=inputs.station_power_mw,
        battery_level=battery_level,
        initial_charge_price=initial_charge_price,
    )
    print(f"  Battery level : {battery_level:.4f} MWh (from {battery_pct:.1f}% total charge)")

    print("Running optimiser...")
    result = optimise_battery_schedule(
        prices=inputs.prices,
        cfg=cfg,
        state=state,
        start_slot=0,
    )

    print_schedule(result, inputs.prices)

    if csv_out:
        write_schedule_csv(csv_out, result, inputs.prices, inputs.time_labels)
        print(f"Schedule written to {csv_out}")

    if xlsx_out:
        generate_formatted_excel(
            template_path=template,
            output_path=xlsx_out,
            schedule=result.schedule,
            prices=inputs.prices,
            threshold=threshold,
            color_map=_color_map(app_cfg),
        )
        print(f"Formatted Excel written to {xlsx_out}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument(
        "excel_path", nargs="?", default=None,
        help="Path to a single input .xlsx file. Omit this and use --input-dir instead "
             "to process a whole folder of files.",
    )
    parser.add_argument(
        "--input-dir", default=None,
        help="Batch mode: process every .xlsx file in this folder (instead of a single "
             "excel_path). The template file, if it lives in this folder, is skipped "
             "automatically.",
    )
    parser.add_argument(
        "--output-dir", default="output",
        help="Batch mode: folder to write outputs into (created if missing). "
             "Each input file 'foo.xlsx' produces 'foo_schedule.csv' and/or "
             "'foo_output.xlsx' here. Default: 'output'.",
    )
    parser.add_argument("--sheet", default="BESS (15)", help="Sheet name to read (default: 'BESS (15)')")
    parser.add_argument(
        "--power-unit", choices=["kw", "mw"], default="kw",
        help="Unit of the value in C3. Default 'kw' (divided by 1000 to get MW). "
             "Use 'mw' if your sheet already stores station power in MW.",
    )
    parser.add_argument(
        "--battery-pct", type=float, default=12.0,
        help="Starting total battery charge %% (same default as calculator.py's main entry point: 12).",
    )
    parser.add_argument(
        "--initial-charge-price", type=float, default=0.0,
        help="Average price the currently-stored energy was charged at (EUR/MWh). Default 0.0 (unknown/most permissive).",
    )
    parser.add_argument(
        "--out", default=None,
        help="Single-file mode only: path to write the schedule as CSV. "
             "(In batch mode, CSV filenames are generated automatically.)",
    )
    parser.add_argument(
        "--xlsx-out", default=None,
        help="Single-file mode only: path to write a formatted .xlsx matching calculator.py's "
             "output. (In batch mode, xlsx filenames are generated automatically whenever "
             "--template is provided.)",
    )
    parser.add_argument(
        "--template", default="excel_template.xlsx",
        help="Path to the real excel_template.xlsx that calculator.py duplicates. "
             "In single-file mode this is only used if --xlsx-out is given. In batch mode, "
             "providing this enables xlsx output for every processed file. "
             "Default: 'excel_template.xlsx' in the cwd.",
    )
    parser.add_argument(
        "--threshold", type=float, default=red_line_threshold,
        help=f"Price threshold (EUR/MWh) below which chart labels are colored red (default {red_line_threshold}).",
    )
    args = parser.parse_args()

    if args.input_dir:
        # ------------------------------------------------------------
        # Batch mode: every .xlsx in --input-dir becomes its own
        # schedule.csv (+ formatted .xlsx if --template resolves)
        # ------------------------------------------------------------
        input_dir = args.input_dir
        output_dir = args.output_dir
        os.makedirs(output_dir, exist_ok=True)

        template_basename = os.path.basename(args.template) if args.template else None
        make_xlsx = bool(args.template) and os.path.isfile(args.template)
        if args.template and not make_xlsx:
            print(f"Note: template '{args.template}' not found — xlsx output will be skipped "
                  f"for all files (CSV output still generated).")

        candidates = sorted(
            f for f in os.listdir(input_dir)
            if f.lower().endswith((".xlsx", ".xlsm")) and not f.startswith("~$")
        )
        # Don't try to process the template itself if it happens to live in input_dir.
        if template_basename:
            candidates = [f for f in candidates if f != template_basename]

        if not candidates:
            print(f"No .xlsx files found in '{input_dir}'.")
            return

        print(f"Found {len(candidates)} file(s) in '{input_dir}':")
        for f in candidates:
            print(f"  - {f}")
        print()

        succeeded, failed = [], []
        for fname in candidates:
            stem = os.path.splitext(fname)[0]
            in_path = os.path.join(input_dir, fname)
            csv_out = os.path.join(output_dir, f"{stem}_schedule.csv")
            xlsx_out = os.path.join(output_dir, f"{stem}_output.xlsx") if make_xlsx else None

            print("=" * 78)
            print(f"Processing '{fname}'")
            print("=" * 78)
            try:
                process_one_file(
                    excel_path=in_path,
                    sheet=args.sheet,
                    power_unit=args.power_unit,
                    battery_pct=args.battery_pct,
                    initial_charge_price=args.initial_charge_price,
                    threshold=args.threshold,
                    template=args.template if make_xlsx else None,
                    csv_out=csv_out,
                    xlsx_out=xlsx_out,
                )
                succeeded.append(fname)
            except Exception as e:
                print(f"FAILED on '{fname}': {e}")
                failed.append((fname, str(e)))
            print()

        print("=" * 78)
        print(f"Batch complete: {len(succeeded)} succeeded, {len(failed)} failed.")
        if failed:
            print("Failures:")
            for fname, err in failed:
                print(f"  - {fname}: {err}")
        print(f"Outputs written to '{output_dir}'.")

    else:
        # ------------------------------------------------------------
        # Single-file mode (original behaviour)
        # ------------------------------------------------------------
        if not args.excel_path:
            parser.error("excel_path is required unless --input-dir is given.")

        process_one_file(
            excel_path=args.excel_path,
            sheet=args.sheet,
            power_unit=args.power_unit,
            battery_pct=args.battery_pct,
            initial_charge_price=args.initial_charge_price,
            threshold=args.threshold,
            template=args.template,
            csv_out=args.out,
            xlsx_out=args.xlsx_out,
        )


if __name__ == "__main__":
    main()
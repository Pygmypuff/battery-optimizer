import copy
from dataclasses import dataclass
from enum import Enum
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.colors import ColorChoice, SchemeColor
import pandas as pd

from battery_optimizer import (
    BatteryAction,
    StationConfig,
    StationState,
    compute_charge_discharge_ratio,
    optimise_battery_schedule,
    print_schedule,
    rerun_for_remaining_day,
)

BASE_CFG = StationConfig(
    max_charge_rate =      0.4, # C   (MW)
    max_sell_rate =        0.5, # S   (MW)
    battery_capacity =     0.77353, # B   (MWh)
    min_price_delta =      40, # Y   (EUR/MWh)
    min_discharge_price =  52.5, # T   (EUR/MWh)
    discharge_loss_pct =   2, # 0–100 (%)
)

# --- Domain types ---

class BatteryAction(str, Enum):
    CHARGE    = "CHARGE"
    DISCHARGE = "DISCHARGE"
    HOLD      = "HOLD"

@dataclass
class SlotResult:
    """Outcome for a single 15-minute slot."""
    slot_index:        int
    action:            BatteryAction
    energy_charged:    float   # MWh added to battery         (0 unless CHARGE)
    energy_discharged: float   # MWh drawn from battery       (0 unless DISCHARGE)
    power_sold:        float   # MW sold to grid this slot
    revenue:           float   # EUR earned this slot
    battery_level_end: float   # MWh in battery at end of slot

COLOR_MAP = {
    BatteryAction.CHARGE:    "4472C4",  # blue
    BatteryAction.DISCHARGE: "ED7D31",  # orange
    BatteryAction.HOLD:      "A6A6A6",  # grey
}


# --- Step 1: Read column C values from Sheet1 starting at C2 ---

def read_column_c(filepath: str) -> list[float]:
    df = pd.read_excel(filepath, sheet_name="Sheet1", header=None)
    values = pd.concat([df.iloc[93:, 3], df.iloc[1:, 2]], ignore_index=True) # Adjusted to read from D94 - first 4 values and then C2 downwards
    return [float(v) for v in values.dropna()]


# RUN BATTERY OPTIMIZER SCRIPT ON THE 96 VALUES TO GET A LIST OF 96 SlotResults
def battery_optimizer_run(prices: list[float]):
    state = StationState(
        station_power = 0.28,   # P = 80 MW
        battery_level = 0,   # starting with 10 MWh stored
    )

    ratio = compute_charge_discharge_ratio(state.station_power, BASE_CFG)

    print("=" * 78)
    print("  Full-day optimisation from 00:00")
    print(f"  P = {state.station_power*1000:.0f} MW  |  "
          f"Computed ratio X = {ratio:.4f}  |  "
          f"Battery = {state.battery_level*1000:.0f} MWh")
    print("=" * 78)

    result = optimise_battery_schedule(prices=prices, cfg=BASE_CFG, state=state)
    return result

# --- Step 2: Copy sheets into a new workbook ---

def _copy_sheet(ws_src, wb_dst, sheet_name: str):
    """Copy a single worksheet into wb_dst."""
    if sheet_name == wb_dst.active.title and len(wb_dst.worksheets) == 1 and wb_dst.active.max_row is None:
        ws_dst = wb_dst.active
        ws_dst.title = sheet_name
    else:
        ws_dst = wb_dst.create_sheet(sheet_name)

    for row in ws_src.iter_rows():
        for cell in row:
            new_cell = ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy.copy(cell.font)
                new_cell.fill = copy.copy(cell.fill)
                new_cell.border = copy.copy(cell.border)
                new_cell.alignment = copy.copy(cell.alignment)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy.copy(cell.protection)

    for col_letter, col_dim in ws_src.column_dimensions.items():
        ws_dst.column_dimensions[col_letter].width = col_dim.width
        ws_dst.column_dimensions[col_letter].hidden = col_dim.hidden

    for row_num, row_dim in ws_src.row_dimensions.items():
        ws_dst.row_dimensions[row_num].height = row_dim.height
        ws_dst.row_dimensions[row_num].hidden = row_dim.hidden

    for merged_range in ws_src.merged_cells.ranges:
        ws_dst.merge_cells(str(merged_range))

    for chart in ws_src._charts:
        ws_dst.add_chart(copy.deepcopy(chart))


def copy_sheets_to_new_workbook(src_path: str, sheet_names: list[str], dst_path: str) -> None:
    """Copy multiple sheets (in order) from src into a new workbook."""
    wb_src = load_workbook(src_path)
    wb_dst = Workbook()

    for sheet_name in sheet_names:
        _copy_sheet(wb_src[sheet_name], wb_dst, sheet_name)

    if "Sheet" in wb_dst.sheetnames and "Sheet" not in sheet_names:
        del wb_dst["Sheet"]

    wb_dst.save(dst_path)


# --- Step 3: Color the chart bars based on slot results ---

def color_chart_bars(filepath: str, slots: list[SlotResult]) -> None:
    if len(slots) != 100:
        raise ValueError(f"Expected 100 SlotResults, got {len(slots)}")

    wb = load_workbook(filepath)
    ws = wb["BESS (15)"]
    chart = ws._charts[0]
    ser = chart.series[0]

    dpt_by_idx = {pt.idx: pt for pt in ser.dPt}

    for slot in slots:
        hex_color = COLOR_MAP[slot.action]
        pt = dpt_by_idx[slot.slot_index]
        pt.spPr.solidFill.schemeClr = None
        pt.spPr.solidFill.srgbClr = hex_color

    wb.save(filepath)


# --- Main pipeline ---

def process_bess(src_path: str, dst_path: str) -> list[float]:
    """
    Full pipeline:
      1. Read 96 float values from Sheet1 column C (starting C2)
      2. Run battery optimizer on these values to get a list of 96 SlotResults
      3. Copy Sheet1 and "BESS (15)" into a new Excel file at dst_path
      4. Color the chart bars according to each SlotResult's BatteryAction

    Returns the list of 96 float values read in step 1.
    """
    print("Reading column C values...")
    values = read_column_c(src_path)
    print(f"  Read {len(values)} values. First 3: {values[:3]}")

    print("Running battery optimizer...")
    result = battery_optimizer_run(values)

    print("Copying 'Sheet1' and 'BESS (15)' to new file...")
    copy_sheets_to_new_workbook(src_path, ["Sheet1", "BESS (15)"], dst_path)
    print(f"  Saved to {dst_path}")

    print("Coloring chart bars...")
    color_chart_bars(dst_path, result.schedule)
    print("  Done.")

    return result


if __name__ == "__main__":

    result = process_bess(
        src_path="cenas(BESS).xlsx",
        dst_path="bess15_output.xlsx",
    )


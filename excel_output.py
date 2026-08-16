"""
excel_output.py
================
Everything that writes/formats the generated output workbook — used to be
duplicated near-verbatim between calculator.py (live Nordpool run) and
test_from_excel.py (test-mode run from a saved input file), since
test_from_excel.py deliberately avoids importing calculator.py (that would
drag in the `nordpool` package at module scope, not needed for offline
testing). Consolidated here instead, following the same pattern as
station_config.py / app_paths.py: no dependency on calculator.py, nordpool,
pytz, or pandas — just openpyxl, zipfile and battery_optimizer.BatteryAction
(itself dependency-light) — so both call sites can share it for free.

generate_formatted_excel() is the one entry point most callers want; the
individual steps (duplicate_excel, replace_column_c, color_chart_bars,
color_label_text_below_threshold, apply_chart_background) are exposed too
since calculator.py's live run interleaves print() progress messages
between them.
"""

from __future__ import annotations

import os
import re
import shutil
import zipfile

from openpyxl import load_workbook
from openpyxl.chart.label import DataLabel
from openpyxl.chart.text import RichText
from openpyxl.drawing.colors import ColorChoice, SchemeColor
from openpyxl.drawing.text import (
    CharacterProperties, ListStyle, Paragraph, ParagraphProperties, RichTextProperties,
)

from battery_optimizer import BatteryAction

TEMPLATE_SHEET = "tabula"
TEMPLATE_FIRST_ROW = 2
TEMPLATE_ROWS = 140  # rows 2..141


def build_color_map(app_cfg) -> dict:
    """Built fresh from config on every run — see station_config.py's
    charge_color/discharge_color/hold_color, editable from the GUI's
    "Customize output chart" window."""
    return {
        BatteryAction.CHARGE:    app_cfg.charge_color,
        BatteryAction.DISCHARGE: app_cfg.discharge_color,
        BatteryAction.HOLD:      app_cfg.hold_color,
    }


def duplicate_excel(src_path: str, dst_path: str) -> None:
    """Duplicate an excel file, preserving all formatting and charts."""
    shutil.copy2(src_path, dst_path)


def replace_column_c(filepath: str, values: list[float]) -> None:
    """Replace values in column C (starting from row TEMPLATE_FIRST_ROW) with the given values."""
    if len(values) != TEMPLATE_ROWS:
        raise ValueError(f"Expected {TEMPLATE_ROWS} values, got {len(values)}")

    wb = load_workbook(filepath)
    ws = wb[TEMPLATE_SHEET]

    for i, value in enumerate(values):
        ws.cell(row=i + TEMPLATE_FIRST_ROW, column=3, value=value)

    wb.save(filepath)


def color_chart_bars(filepath: str, slots, start_index: int, color_map: dict) -> None:
    """Colors the bars in the excel chart to match the slot actions."""
    if len(slots) + start_index != TEMPLATE_ROWS:
        raise ValueError(f"Expected {TEMPLATE_ROWS - start_index} SlotResults, got {len(slots)}")

    wb = load_workbook(filepath)
    ws = wb[TEMPLATE_SHEET]
    chart = ws._charts[0]
    ser = chart.series[0]

    dpt_by_idx = {pt.idx: pt for pt in ser.dPt}

    for i, slot in enumerate(slots):
        bar_index = start_index + i
        pt = dpt_by_idx[bar_index]
        pt.spPr.solidFill.schemeClr = None
        pt.spPr.solidFill.srgbClr = color_map[slot.action]

    wb.save(filepath)


def _make_txPr(hex_color: str) -> RichText:
    """RichText run properties used to color a chart data-label's text."""
    rPr = CharacterProperties(sz=1000.0, b=True)
    rPr.solidFill = ColorChoice()
    rPr.solidFill.srgbClr = hex_color
    pPr = ParagraphProperties(defRPr=rPr)
    para = Paragraph(pPr=pPr, endParaRPr=CharacterProperties())
    bodyPr = RichTextProperties(rot=-5400000, anchor="ctr", anchorCtr=True)
    return RichText(bodyPr=bodyPr, lstStyle=ListStyle(), p=[para])


def _make_txPr_default() -> RichText:
    solidFill = ColorChoice()
    solidFill.schemeClr = SchemeColor(val="tx1")
    rPr = CharacterProperties(sz=1000.0, b=True)
    rPr.solidFill = solidFill
    pPr = ParagraphProperties(defRPr=rPr)
    para = Paragraph(pPr=pPr, endParaRPr=CharacterProperties())
    bodyPr = RichTextProperties(rot=-5400000, anchor="ctr", anchorCtr=True)
    return RichText(bodyPr=bodyPr, lstStyle=ListStyle(), p=[para])


def color_label_text_below_threshold(
    filepath: str,
    values: list[float],
    threshold: float,
    sheet_name: str = TEMPLATE_SHEET,
    chart_index: int = 0,
    series_index: int = 0,
) -> None:
    """Colors the data labels red if they are below threshold."""
    if len(values) != TEMPLATE_ROWS:
        raise ValueError(f"Expected {TEMPLATE_ROWS} values, got {len(values)}")

    wb = load_workbook(filepath)
    ws = wb[sheet_name]
    chart = ws._charts[chart_index]
    ser = chart.series[series_index]

    # ser.dLbls.dLbl only holds *overrides* for points someone has already
    # customized — the template doesn't have one for every point, so
    # iterating it directly silently skips whichever indices happen to be
    # missing (they just keep whatever this point's inherited/default
    # coloring is, red or not, regardless of `values`). Every one of the
    # TEMPLATE_ROWS points needs an explicit color decision, so build any
    # missing entries instead of only touching the ones that already exist.
    existing = {lbl.idx: lbl for lbl in ser.dLbls.dLbl}
    for i, value in enumerate(values):
        txPr = _make_txPr("FF0000") if value < threshold else _make_txPr_default()
        if i in existing:
            existing[i].txPr = txPr
        else:
            ser.dLbls.dLbl.append(DataLabel(
                idx=i,
                txPr=txPr,
                showLegendKey=False,
                showVal=True,
                showCatName=False,
                showSerName=False,
                showPercent=False,
                showBubbleSize=False,
            ))

    wb.save(filepath)


def apply_chart_background(filepath: str) -> None:
    """
    Injects a gradient background into the chart XML to visually distinguish
    between 24h segments. Call this AFTER all openpyxl wb.save() calls.
    """
    s1_end = round(40 / TEMPLATE_ROWS * 100_000)
    s2_end = round(136 / TEMPLATE_ROWS * 100_000)

    files: dict[str, bytes] = {}
    with zipfile.ZipFile(filepath, "r") as z:
        for name in z.namelist():
            files[name] = z.read(name)

    chart_xml = files["xl/charts/chart1.xml"].decode("utf-8")

    # openpyxl strips namespace prefixes on save (e.g. <c:plotArea> → <plotArea>)
    # so we detect which form is present and adapt accordingly.
    if "</c:plotArea>" in chart_xml:
        close_tag = "</c:plotArea>"
        open_spPr = (
            '<c:spPr'
            ' xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart"'
            ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        )
        close_spPr = "</c:spPr>"
    else:
        close_tag = "</plotArea>"
        open_spPr = '<spPr xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        close_spPr = "</spPr>"

    gradient_spPr = (
        open_spPr
        + "<a:gradFill><a:gsLst>"
        + '<a:gs pos="0"><a:srgbClr val="FFFFFF"/></a:gs>'
        + f'<a:gs pos="{s1_end}"><a:srgbClr val="FFFFFF"/></a:gs>'
        + f'<a:gs pos="{s1_end + 1}"><a:srgbClr val="E0E0E0"/></a:gs>'
        + f'<a:gs pos="{s2_end}"><a:srgbClr val="E0E0E0"/></a:gs>'
        + f'<a:gs pos="{s2_end + 1}"><a:srgbClr val="FFFFFF"/></a:gs>'
        + '<a:gs pos="100000"><a:srgbClr val="FFFFFF"/></a:gs>'
        + "</a:gsLst>"
        + '<a:lin ang="0" scaled="0"/>'
        + "</a:gradFill>"
        + "<a:ln><a:noFill/></a:ln>"
        + close_spPr
    )

    # Remove any previously injected plotArea spPr (makes this call idempotent)
    for open_pat, close_pat in [
        (r"<c:spPr[^>]*>", r"</c:spPr>\s*</c:plotArea>"),
        (r"<spPr[^>]*>",   r"</spPr>\s*</plotArea>"),
    ]:
        chart_xml = re.sub(
            open_pat + r".*?" + close_pat,
            close_tag,
            chart_xml,
            count=1,
            flags=re.DOTALL,
        )

    # Inject the gradient spPr immediately before </plotArea>
    chart_xml = chart_xml.replace(close_tag, gradient_spPr + close_tag, 1)

    files["xl/charts/chart1.xml"] = chart_xml.encode("utf-8")

    # Write back atomically
    tmp = filepath + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in files.items():
            zout.writestr(name, data)
    os.replace(tmp, filepath)


def generate_formatted_excel(
    template_path: str,
    output_path: str,
    schedule,
    prices: list[float],
    threshold: float,
    color_map: dict,
    start_index: int = 0,
) -> None:
    """
    Duplicates `template_path` to `output_path` and fills it in: prices,
    chart bar colors by action, red price labels below `threshold`, and the
    24h-segment background band — the full pipeline both calculator.py's
    live run and test_from_excel.py's test-mode run need.

    `start_index` is 0 for a full-day schedule (test mode always starts at
    slot 0); calculator.py's live run passes the actual slot the schedule
    starts at, since a 14:00 (or later) run only covers the remainder of
    the day.
    """
    print(f"Duplicating template '{template_path}' -> '{output_path}'...")
    duplicate_excel(template_path, output_path)

    print("Writing prices into template...")
    replace_column_c(output_path, prices)

    print("Coloring chart bars by action...")
    color_chart_bars(output_path, schedule, start_index=start_index, color_map=color_map)

    print("Coloring price labels below threshold...")
    color_label_text_below_threshold(output_path, prices, threshold)

    print("Applying chart background bands...")
    apply_chart_background(output_path)  # must be called last
